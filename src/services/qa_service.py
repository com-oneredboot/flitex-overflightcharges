"""QA service layer for route string testing harness.

Business logic for flight plan import, test run execution, review,
comparison, and export operations. All QA data lives in the dedicated
`qa` PostgreSQL schema.

Validates Requirements: 5.1, 5.2, 5.3, 5.4, 5.5
"""

import csv
import io
import json
import logging
import subprocess
import time
import uuid
from typing import BinaryIO

from openpyxl import load_workbook
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from src.models.qa import QAFlightPlan, QATestRun, QATestRunResult, QATestRunReview
from src.schemas.qa import ComparisonCategory, FlightPlanImportResponse
from src.services.route_parser import RouteParser


logger = logging.getLogger(__name__)

# Columns that map directly from Excel header to DB column name
DIRECT_COLUMN_MAP = {
    "scheduled_departure_dtmz": "scheduled_departure_dtmz",
    "departure_icao_aerodrome_code": "departure_icao_aerodrome_code",
    "operational_icao_carrier_code": "operational_icao_carrier_code",
    "flight_number": "flight_number",
    "release_number": "release_number",
    "aircraft_type": "aircraft_type",
    "icao_route": "icao_route",
    "destination_aerodrome_code": "destination_aerodrome_code",
    "total_estimated_elapsed_time": "total_estimated_elapsed_time",
    "alternate_aerodrome_list": "alternate_aerodrome_list",
    "hash_code": "hash_code",
}

# Known column name variants that map to a different DB column
COLUMN_VARIANTS = {
    "main_route_details": "icao_route",
}

# Columns that must be present for a valid import
REQUIRED_COLUMNS = {"icao_route", "hash_code"}


def _build_column_mapping(headers: list[str]) -> dict[int, str]:
    """Map Excel column indexes to DB column names using case-insensitive matching.

    Handles direct column name matches and known variants (e.g.
    ``main_route_details`` → ``icao_route``).

    Args:
        headers: List of header strings from row 1 of the Excel file.

    Returns:
        Dict mapping column index (0-based) to the target DB column name.
    """
    mapping: dict[int, str] = {}
    for idx, header in enumerate(headers):
        if header is None:
            continue
        normalized = str(header).strip().lower()
        if normalized in DIRECT_COLUMN_MAP:
            mapping[idx] = DIRECT_COLUMN_MAP[normalized]
        elif normalized in COLUMN_VARIANTS:
            mapping[idx] = COLUMN_VARIANTS[normalized]
    return mapping


def _validate_required_columns(mapping: dict[int, str]) -> list[str]:
    """Return list of required DB columns missing from the mapping.

    The check accounts for variant columns — e.g. if ``main_route_details``
    is present it satisfies the ``icao_route`` requirement.

    Args:
        mapping: Column index → DB column name mapping.

    Returns:
        Sorted list of missing required column names (empty if all present).
    """
    mapped_db_columns = set(mapping.values())
    missing = REQUIRED_COLUMNS - mapped_db_columns
    return sorted(missing)


def import_flight_plans(
    file: BinaryIO,
    db: Session,
    source_filename: str = "",
) -> FlightPlanImportResponse:
    """Import ICAO flight plan data from an Excel file into qa.flight_plans.

    Reads the uploaded ``.xlsx`` bytes, maps columns to DB fields
    (case-insensitive, handles known variants), validates that required
    columns are present, and inserts rows that don't already exist
    (matched by ``hash_code``).

    Args:
        file: File-like object containing the Excel bytes.
        db: SQLAlchemy database session.
        source_filename: Original filename for the ``source_file`` column.

    Returns:
        FlightPlanImportResponse with total, imported, and skipped counts.

    Raises:
        ValueError: If the file is not a valid ``.xlsx``, contains no data
            rows, or is missing required columns.

    Validates Requirements: 5.1, 5.2, 5.3, 5.4, 5.5
    """
    # --- Load workbook -------------------------------------------------------
    try:
        contents = file.read()
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Invalid file format. Expected .xlsx") from exc

    ws = wb.active
    if ws is None:
        raise ValueError("Invalid file format. Expected .xlsx")

    # --- Extract headers from row 1 -----------------------------------------
    rows = ws.iter_rows()
    try:
        header_row = next(rows)
    except StopIteration:
        raise ValueError("File contains no data rows")

    headers = [cell.value for cell in header_row]
    column_mapping = _build_column_mapping(headers)

    # --- Validate required columns -------------------------------------------
    missing = _validate_required_columns(column_mapping)
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")

    # --- Pre-fetch existing hash_codes for fast duplicate detection -----------
    existing_hashes: set[str] = set()
    hash_col_idx: int | None = None
    for idx, db_col in column_mapping.items():
        if db_col == "hash_code":
            hash_col_idx = idx
            break

    if hash_col_idx is None:
        raise ValueError("Missing columns: hash_code")

    existing_rows = db.query(QAFlightPlan.hash_code).all()
    existing_hashes = {row[0] for row in existing_rows if row[0] is not None}

    # --- Process data rows ---------------------------------------------------
    total = 0
    imported = 0
    skipped = 0

    for row in rows:
        cell_values = [cell.value for cell in row]

        # Skip completely empty rows
        if all(v is None for v in cell_values):
            continue

        total += 1

        # Build a dict of DB column → value for this row
        record_data: dict = {}
        for col_idx, db_col in column_mapping.items():
            value = cell_values[col_idx] if col_idx < len(cell_values) else None
            record_data[db_col] = value

        hash_value = record_data.get("hash_code")
        if hash_value is not None:
            hash_value = str(hash_value)
            record_data["hash_code"] = hash_value

        # Skip duplicates (Requirement 5.3)
        if hash_value and hash_value in existing_hashes:
            skipped += 1
            continue

        # Convert release_number to int if present
        rn = record_data.get("release_number")
        if rn is not None:
            try:
                record_data["release_number"] = int(rn)
            except (TypeError, ValueError):
                record_data["release_number"] = None

        record_data["source_file"] = source_filename

        flight_plan = QAFlightPlan(**record_data)
        db.add(flight_plan)

        if hash_value:
            existing_hashes.add(hash_value)

        imported += 1

    db.commit()
    wb.close()

    logger.info(
        "Flight plan import complete: total=%d imported=%d skipped=%d file=%s",
        total,
        imported,
        skipped,
        source_filename,
    )

    return FlightPlanImportResponse(
        total_rows=total,
        imported=imported,
        skipped=skipped,
    )


def _get_git_commit_sha() -> str:
    """Get the current git commit SHA via subprocess.

    Falls back to ``"unknown"`` if git is unavailable or the command fails.

    Returns:
        The 40-character commit SHA string, or ``"unknown"``.

    Validates Requirement: 6.1
    """
    try:
        result = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            capture_output=True, text=True, timeout=5
        )
        return result.stdout.strip() if result.returncode == 0 else "unknown"
    except Exception:
        return "unknown"


def _compute_fir_boundary_hash(db: Session) -> str | None:
    """Compute an MD5 hash of the current FIR boundary dataset.

    Queries ``reference.fir_boundaries`` and produces a deterministic hash
    from the ordered list of boundary IDs. Returns ``None`` if the table
    is unavailable or the query fails.

    Args:
        db: SQLAlchemy database session.

    Returns:
        MD5 hex string or ``None``.

    Validates Requirement: 2.1
    """
    try:
        result = db.execute(text(
            "SELECT MD5(string_agg(id::text, ',' ORDER BY id)) FROM reference.fir_boundaries"
        )).scalar()
        return result
    except Exception:
        return None


def create_test_run(
    notes: str | None,
    created_by: str | None,
    db: Session,
) -> QATestRun:
    """Create a new QA test run record with pending status.

    Captures the current git commit SHA and FIR boundary hash for
    traceability. The run is created with status ``"pending"`` and
    zero counts.

    Args:
        notes: Optional free-text notes for this run.
        created_by: Optional identifier of the person who triggered the run.
        db: SQLAlchemy database session.

    Returns:
        The newly created QATestRun record.

    Validates Requirements: 6.1
    """
    commit_sha = _get_git_commit_sha()
    fir_boundary_hash = _compute_fir_boundary_hash(db)

    test_run = QATestRun(
        id=uuid.uuid4(),
        commit_sha=commit_sha,
        fir_boundary_hash=fir_boundary_hash,
        notes=notes,
        created_by=created_by,
        status="pending",
        total_flight_plans=0,
        completed_count=0,
        failed_count=0,
    )
    db.add(test_run)
    db.commit()
    db.refresh(test_run)

    logger.info(
        "Created test run %s (commit=%s, fir_hash=%s)",
        test_run.id,
        commit_sha,
        fir_boundary_hash,
    )

    return test_run


def _compute_health_status(
    resolved_count: int,
    unresolved_count: int,
    error_message: str | None,
) -> str:
    """Compute the health status for a test run result.

    Logic mirrors the frontend getRouteHealth():
    - fail: error present, or 0 resolved, or >20% unresolved
    - pass: no error, resolved > 0, 0 unresolved
    - warning: no error, resolved > 0, unresolved ratio <= 20%

    Args:
        resolved_count: Number of resolved waypoints.
        unresolved_count: Number of unresolved tokens.
        error_message: Error message from the parser, or None.

    Returns:
        One of 'pass', 'warning', 'fail'.
    """
    if error_message:
        return "fail"
    if resolved_count == 0:
        return "fail"
    if unresolved_count == 0:
        return "pass"
    ratio = unresolved_count / (resolved_count + unresolved_count)
    return "fail" if ratio > 0.2 else "warning"


def execute_test_run(run_id: str, db: Session) -> QATestRun:
    """Execute a QA test run against all flight plans in the store.

    Sets the run status to ``"running"``, iterates every flight plan in
    ``qa.flight_plans``, calls ``RouteParser.parse_route()`` and
    ``identify_fir_crossings_db()`` for each, and stores a
    ``QATestRunResult`` with the structured output.

    Each flight plan is wrapped in try/except so that individual parser
    failures are captured in ``error_message`` without aborting the run.

    On completion the run status is set to ``"completed"`` with final
    counts. If a critical error prevents the loop from finishing the
    status is set to ``"failed"``.

    Args:
        run_id: UUID of the test run to execute.
        db: SQLAlchemy database session.

    Returns:
        The updated QATestRun record.

    Raises:
        ValueError: If the run_id does not exist.

    Validates Requirements: 6.2, 6.3, 6.4, 6.5, 6.6, 6.7, 6.8
    """
    # 1. Fetch the test run record
    test_run = db.query(QATestRun).filter(QATestRun.id == run_id).first()
    if not test_run:
        raise ValueError(f"Test run {run_id} not found")

    # 2. Set status to "running"
    test_run.status = "running"
    db.commit()

    # 3. Get all flight plans
    flight_plans = db.query(QAFlightPlan).all()
    test_run.total_flight_plans = len(flight_plans)
    db.commit()

    completed_count = 0
    failed_count = 0
    parser = RouteParser()

    try:
        # 4. Process each flight plan
        for fp in flight_plans:
            start_time = time.perf_counter()
            try:
                # Validate route string (lightweight, no flight plan context)
                token_result = parser.validate_route_string(
                    fp.icao_route or "",
                    db,
                )

                waypoints = token_result.resolved_waypoints

                # Call identify_fir_crossings_db with the resolved waypoints
                fir_crossings = parser.identify_fir_crossings_db(waypoints, db)

                elapsed_ms = int((time.perf_counter() - start_time) * 1000)

                # Build result dicts from parser output
                resolved_waypoints_data = [
                    {
                        "identifier": wp.identifier,
                        "latitude": wp.latitude,
                        "longitude": wp.longitude,
                        "source_table": wp.source_table,
                    }
                    for wp in waypoints
                ]

                fir_crossings_data = [
                    {
                        "icao_code": fc.icao_code,
                        "fir_name": fc.fir_name,
                        "sequence": idx + 1,
                    }
                    for idx, fc in enumerate(fir_crossings)
                ]

                # Get unresolved tokens directly from TokenResolutionResult
                unresolved_tokens_data = [
                    {"token": tr.raw, "classification": tr.classification}
                    for tr in token_result.unresolved_tokens
                ]

                result = QATestRunResult(
                    id=uuid.uuid4(),
                    test_run_id=test_run.id,
                    flight_plan_id=fp.id,
                    resolved_waypoints=resolved_waypoints_data,
                    fir_crossings=fir_crossings_data,
                    unresolved_tokens=unresolved_tokens_data,
                    parse_duration_ms=elapsed_ms,
                    error_message=None,
                    health_status=_compute_health_status(
                        len(resolved_waypoints_data),
                        len(unresolved_tokens_data),
                        None,
                    ),
                )
                db.add(result)
                completed_count += 1

            except Exception as exc:
                elapsed_ms = int((time.perf_counter() - start_time) * 1000)

                result = QATestRunResult(
                    id=uuid.uuid4(),
                    test_run_id=test_run.id,
                    flight_plan_id=fp.id,
                    resolved_waypoints=None,
                    fir_crossings=None,
                    unresolved_tokens=None,
                    parse_duration_ms=elapsed_ms,
                    error_message=str(exc),
                    health_status="fail",
                )
                db.add(result)
                failed_count += 1

                logger.warning(
                    "Flight plan %d failed in run %s: %s",
                    fp.id,
                    test_run.id,
                    str(exc),
                )

        # 5. Update final counts and set status to "completed"
        test_run.completed_count = completed_count
        test_run.failed_count = failed_count
        test_run.status = "completed"
        db.commit()
        db.refresh(test_run)

        logger.info(
            "Test run %s completed: total=%d completed=%d failed=%d",
            test_run.id,
            test_run.total_flight_plans,
            completed_count,
            failed_count,
        )

    except Exception as exc:
        # Critical error — mark run as failed
        test_run.completed_count = completed_count
        test_run.failed_count = failed_count
        test_run.status = "failed"
        db.commit()
        db.refresh(test_run)

        logger.error(
            "Test run %s failed critically: %s",
            test_run.id,
            str(exc),
            exc_info=True,
        )

    return test_run


def get_test_runs(page: int, page_size: int, db: Session) -> dict:
    """Return a paginated list of test runs ordered by run_timestamp descending.

    For each run, computes the review progress: how many results have at
    least one review vs the total number of results.

    Args:
        page: 1-based page number.
        page_size: Number of items per page.
        db: SQLAlchemy database session.

    Returns:
        Dict with ``items`` (list of TestRunResponse-compatible dicts),
        ``total``, ``page``, and ``page_size``.

    Validates Requirements: 7.1, 7.2
    """
    total = db.query(func.count(QATestRun.id)).scalar() or 0

    offset = (page - 1) * page_size
    runs = (
        db.query(QATestRun)
        .order_by(QATestRun.run_timestamp.desc())
        .offset(offset)
        .limit(page_size)
        .all()
    )

    items = []
    for run in runs:
        total_results = (
            db.query(func.count(QATestRunResult.id))
            .filter(QATestRunResult.test_run_id == run.id)
            .scalar()
        ) or 0

        # Count results that have at least one review
        reviewed_count = (
            db.query(func.count(func.distinct(QATestRunReview.test_run_result_id)))
            .join(QATestRunResult, QATestRunReview.test_run_result_id == QATestRunResult.id)
            .filter(QATestRunResult.test_run_id == run.id)
            .scalar()
        ) or 0

        # Health status breakdown
        health_rows = (
            db.query(QATestRunResult.health_status, func.count(QATestRunResult.id))
            .filter(QATestRunResult.test_run_id == run.id)
            .group_by(QATestRunResult.health_status)
            .all()
        )
        health_counts = {row[0]: row[1] for row in health_rows}

        items.append({
            "id": str(run.id),
            "commit_sha": run.commit_sha or "",
            "run_timestamp": run.run_timestamp,
            "status": run.status or "pending",
            "total_flight_plans": run.total_flight_plans or 0,
            "completed_count": run.completed_count or 0,
            "failed_count": run.failed_count or 0,
            "notes": run.notes,
            "fir_boundary_hash": run.fir_boundary_hash,
            "created_by": run.created_by,
            "reviewed_count": reviewed_count,
            "total_results": total_results,
            "pass_count": health_counts.get("pass", 0),
            "warning_count": health_counts.get("warning", 0),
            "fail_count": health_counts.get("fail", 0),
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


def get_test_run_detail(
    run_id: str,
    db: Session,
    verdict: str | None = None,
    has_error: bool | None = None,
    departure: str | None = None,
    destination: str | None = None,
    health_status: str | None = None,
    page: int = 1,
    page_size: int = 50,
) -> dict:
    """Return detailed results for a specific test run with optional filters.

    Joins results with flight plans and includes the latest review verdict
    for each result. Supports filtering by verdict, error status, and
    departure/destination airport.

    Args:
        run_id: UUID of the test run.
        db: SQLAlchemy database session.
        verdict: Filter by latest review verdict (correct/incorrect/
            needs_investigation/unreviewed).
        has_error: Filter by error presence (True = has error, False = no error).
        departure: Filter by flight plan departure ICAO code.
        destination: Filter by flight plan destination ICAO code.
        page: 1-based page number.
        page_size: Number of results per page.

    Returns:
        Dict with ``run`` (TestRunResponse-compatible dict), ``results``
        (list of TestRunResultResponse-compatible dicts), ``total_results``,
        ``page``, and ``page_size``.

    Raises:
        ValueError: If the run_id does not exist.

    Validates Requirements: 7.1, 7.2, 7.3, 11.6
    """
    # Fetch the test run
    test_run = db.query(QATestRun).filter(QATestRun.id == run_id).first()
    if not test_run:
        raise ValueError(f"Test run {run_id} not found")

    # Build run-level review stats
    total_results_count = (
        db.query(func.count(QATestRunResult.id))
        .filter(QATestRunResult.test_run_id == test_run.id)
        .scalar()
    ) or 0

    reviewed_count = (
        db.query(func.count(func.distinct(QATestRunReview.test_run_result_id)))
        .join(QATestRunResult, QATestRunReview.test_run_result_id == QATestRunResult.id)
        .filter(QATestRunResult.test_run_id == test_run.id)
        .scalar()
    ) or 0

    run_data = {
        "id": str(test_run.id),
        "commit_sha": test_run.commit_sha or "",
        "run_timestamp": test_run.run_timestamp,
        "status": test_run.status or "pending",
        "total_flight_plans": test_run.total_flight_plans or 0,
        "completed_count": test_run.completed_count or 0,
        "failed_count": test_run.failed_count or 0,
        "notes": test_run.notes,
        "fir_boundary_hash": test_run.fir_boundary_hash,
        "created_by": test_run.created_by,
        "reviewed_count": reviewed_count,
        "total_results": total_results_count,
    }

    # Subquery: latest review per result (most recent reviewed_at)
    latest_review_sq = (
        db.query(
            QATestRunReview.test_run_result_id,
            QATestRunReview.verdict,
            QATestRunReview.reviewer_notes,
            func.row_number()
            .over(
                partition_by=QATestRunReview.test_run_result_id,
                order_by=QATestRunReview.reviewed_at.desc(),
            )
            .label("rn"),
        )
        .subquery("latest_review")
    )

    # Base query: results joined with flight plans and latest review
    query = (
        db.query(
            QATestRunResult,
            QAFlightPlan,
            latest_review_sq.c.verdict.label("latest_verdict"),
            latest_review_sq.c.reviewer_notes.label("latest_reviewer_notes"),
        )
        .join(QAFlightPlan, QATestRunResult.flight_plan_id == QAFlightPlan.id)
        .outerjoin(
            latest_review_sq,
            (latest_review_sq.c.test_run_result_id == QATestRunResult.id)
            & (latest_review_sq.c.rn == 1),
        )
        .filter(QATestRunResult.test_run_id == test_run.id)
    )

    # Apply filters
    if has_error is True:
        query = query.filter(QATestRunResult.error_message.isnot(None))
    elif has_error is False:
        query = query.filter(QATestRunResult.error_message.is_(None))

    if departure:
        query = query.filter(
            QAFlightPlan.departure_icao_aerodrome_code == departure
        )

    if destination:
        query = query.filter(
            QAFlightPlan.destination_aerodrome_code == destination
        )

    if verdict:
        if verdict == "unreviewed":
            query = query.filter(latest_review_sq.c.verdict.is_(None))
        else:
            query = query.filter(latest_review_sq.c.verdict == verdict)

    # Health status filter — uses the pre-computed DB column
    if health_status:
        query = query.filter(QATestRunResult.health_status == health_status)

    # Get total count for pagination (after filters)
    total_filtered = query.count()

    # Paginate
    offset = (page - 1) * page_size
    rows = query.offset(offset).limit(page_size).all()

    results = []
    for result, fp, lat_verdict, lat_notes in rows:
        results.append({
            "id": str(result.id),
            "flight_plan_id": fp.id,
            "departure_icao_aerodrome_code": fp.departure_icao_aerodrome_code,
            "destination_aerodrome_code": fp.destination_aerodrome_code,
            "operational_icao_carrier_code": fp.operational_icao_carrier_code,
            "flight_number": fp.flight_number,
            "icao_route": fp.icao_route,
            "resolved_waypoints": result.resolved_waypoints,
            "fir_crossings": result.fir_crossings,
            "unresolved_tokens": result.unresolved_tokens,
            "parse_duration_ms": result.parse_duration_ms,
            "error_message": result.error_message,
            "health_status": result.health_status,
            "latest_verdict": lat_verdict,
            "latest_reviewer_notes": lat_notes,
        })

    return {
        "run": run_data,
        "results": results,
        "total_results": total_filtered,
        "page": page,
        "page_size": page_size,
    }


def get_flight_plans(page: int, page_size: int, db: Session) -> dict:
    """Return a paginated list of flight plans ordered by id.

    Args:
        page: 1-based page number.
        page_size: Number of items per page.
        db: SQLAlchemy database session.

    Returns:
        Dict with ``items`` (list of FlightPlanResponse-compatible dicts),
        ``total``, ``page``, and ``page_size``.

    Validates Requirement: 5.1
    """
    total = db.query(func.count(QAFlightPlan.id)).scalar() or 0

    offset = (page - 1) * page_size
    flight_plans = (
        db.query(QAFlightPlan)
        .order_by(QAFlightPlan.id)
        .offset(offset)
        .limit(page_size)
        .all()
    )

    return {
        "items": flight_plans,
        "total": total,
        "page": page,
        "page_size": page_size,
    }

def submit_review(
    result_id: str,
    verdict: str,
    reviewer_notes: str | None,
    reviewed_by: str,
    db: Session,
) -> QATestRunReview:
    """Create a new review record for a test run result.

    Each call creates a new ``QATestRunReview`` row, preserving the full
    review history.  The latest review is determined by ``reviewed_at``
    timestamp when displaying results.

    Args:
        result_id: UUID of the test run result to review.
        verdict: One of ``correct``, ``incorrect``, ``needs_investigation``.
        reviewer_notes: Optional free-text notes from the reviewer.
        reviewed_by: Identifier of the person submitting the review.
        db: SQLAlchemy database session.

    Returns:
        The newly created ``QATestRunReview`` record.

    Raises:
        ValueError: If the ``result_id`` does not match an existing result.

    Validates Requirements: 4.4, 8.1, 8.2
    """
    result = (
        db.query(QATestRunResult)
        .filter(QATestRunResult.id == result_id)
        .first()
    )
    if not result:
        raise ValueError(f"Test run result not found: {result_id}")

    review = QATestRunReview(
        test_run_result_id=result_id,
        verdict=verdict,
        reviewer_notes=reviewer_notes,
        reviewed_by=reviewed_by,
    )
    db.add(review)
    db.commit()
    db.refresh(review)
    return review


def get_review_history(
    result_id: str,
    db: Session,
) -> list[QATestRunReview]:
    """Return all review records for a test run result, newest first.

    Args:
        result_id: UUID of the test run result.
        db: SQLAlchemy database session.

    Returns:
        List of ``QATestRunReview`` records ordered by ``reviewed_at`` descending.
    """
    return (
        db.query(QATestRunReview)
        .filter(QATestRunReview.test_run_result_id == result_id)
        .order_by(QATestRunReview.reviewed_at.desc())
        .all()
    )


def _categorize_result_pair(
    result_1: QATestRunResult,
    result_2: QATestRunResult,
) -> tuple[ComparisonCategory, dict | None]:
    """Categorize a pair of results and build diff details if non-unchanged.

    Comparison logic per design:
    - unchanged: resolved_waypoints and fir_crossings are identical (deep JSON equality)
    - improved: fewer unresolved tokens OR more FIR crossings in the newer run (run 2)
    - regressed: more unresolved tokens OR fewer FIR crossings in the newer run (run 2)
    - changed: different results that don't clearly fit improved or regressed

    Returns:
        Tuple of (category, diff_details). diff_details is None for unchanged.

    Validates Requirements: 10.2, 10.4
    """
    wp_1 = result_1.resolved_waypoints or []
    wp_2 = result_2.resolved_waypoints or []
    fir_1 = result_1.fir_crossings or []
    fir_2 = result_2.fir_crossings or []
    unresolved_1 = result_1.unresolved_tokens or []
    unresolved_2 = result_2.unresolved_tokens or []

    wp_same = wp_1 == wp_2
    fir_same = fir_1 == fir_2

    if wp_same and fir_same:
        return ComparisonCategory.unchanged, None

    # Build diff details for non-unchanged items
    diff_details: dict = {}

    if not wp_same:
        diff_details["resolved_waypoints"] = {
            "run_1": wp_1,
            "run_2": wp_2,
        }

    if not fir_same:
        diff_details["fir_crossings"] = {
            "run_1": fir_1,
            "run_2": fir_2,
        }

    if unresolved_1 != unresolved_2:
        diff_details["unresolved_tokens"] = {
            "run_1": unresolved_1,
            "run_2": unresolved_2,
        }

    # Determine direction of change
    unresolved_count_1 = len(unresolved_1)
    unresolved_count_2 = len(unresolved_2)
    fir_count_1 = len(fir_1)
    fir_count_2 = len(fir_2)

    fewer_unresolved = unresolved_count_2 < unresolved_count_1
    more_unresolved = unresolved_count_2 > unresolved_count_1
    more_fir = fir_count_2 > fir_count_1
    fewer_fir = fir_count_2 < fir_count_1

    has_improvement_signal = fewer_unresolved or more_fir
    has_regression_signal = more_unresolved or fewer_fir

    if has_improvement_signal and not has_regression_signal:
        return ComparisonCategory.improved, diff_details
    elif has_regression_signal and not has_improvement_signal:
        return ComparisonCategory.regressed, diff_details
    else:
        return ComparisonCategory.changed, diff_details


def compare_runs(run_id_1: str, run_id_2: str, db: Session) -> dict:
    """Compare two test runs and categorize differences per flight plan.

    Joins results from both runs on ``flight_plan_id``, categorizes each
    matched pair as unchanged, improved, regressed, or changed, and
    returns summary counts with diff details for non-unchanged items.

    Args:
        run_id_1: UUID of the first (older/baseline) test run.
        run_id_2: UUID of the second (newer) test run.
        db: SQLAlchemy database session.

    Returns:
        Dict with ``run_1_id``, ``run_2_id``, ``summary`` counts, and
        ``items`` list of per-flight-plan comparison details.

    Raises:
        ValueError: If either run ID does not exist.

    Validates Requirements: 10.1, 10.2, 10.3, 10.4
    """
    # 1. Verify both test runs exist
    run_1 = db.query(QATestRun).filter(QATestRun.id == run_id_1).first()
    if not run_1:
        raise ValueError(f"Test run {run_id_1} not found")

    run_2 = db.query(QATestRun).filter(QATestRun.id == run_id_2).first()
    if not run_2:
        raise ValueError(f"Test run {run_id_2} not found")

    # 2. Query results for both runs, keyed by flight_plan_id
    results_1 = (
        db.query(QATestRunResult)
        .filter(QATestRunResult.test_run_id == run_id_1)
        .all()
    )
    results_2 = (
        db.query(QATestRunResult)
        .filter(QATestRunResult.test_run_id == run_id_2)
        .all()
    )

    map_1 = {r.flight_plan_id: r for r in results_1}
    map_2 = {r.flight_plan_id: r for r in results_2}

    # 3. Only compare flight plans that appear in both runs
    common_fp_ids = set(map_1.keys()) & set(map_2.keys())

    # 4. Categorize each matched pair
    summary = {
        "unchanged": 0,
        "improved": 0,
        "regressed": 0,
        "changed": 0,
    }
    items: list[dict] = []

    for fp_id in sorted(common_fp_ids):
        r1 = map_1[fp_id]
        r2 = map_2[fp_id]

        category, diff_details = _categorize_result_pair(r1, r2)
        summary[category.value] += 1

        # Load flight plan data for context
        flight_plan = (
            db.query(QAFlightPlan).filter(QAFlightPlan.id == fp_id).first()
        )

        item = {
            "flight_plan_id": fp_id,
            "departure_icao_aerodrome_code": (
                flight_plan.departure_icao_aerodrome_code if flight_plan else None
            ),
            "destination_aerodrome_code": (
                flight_plan.destination_aerodrome_code if flight_plan else None
            ),
            "icao_route": flight_plan.icao_route if flight_plan else None,
            "category": category.value,
            "diff_details": diff_details,
        }
        items.append(item)

    return {
        "run_1_id": str(run_id_1),
        "run_2_id": str(run_id_2),
        "summary": summary,
        "items": items,
    }


# ---------------------------------------------------------------------------
# Delete test run
# ---------------------------------------------------------------------------


def delete_test_run(run_id: str, db: Session) -> None:
    """Delete a test run and all associated results and reviews.

    Cascades deletion: reviews → results → test run.

    Args:
        run_id: UUID of the test run to delete.
        db: SQLAlchemy database session.

    Raises:
        ValueError: If the run_id does not exist.
    """
    test_run = db.query(QATestRun).filter(QATestRun.id == run_id).first()
    if not test_run:
        raise ValueError(f"Test run {run_id} not found")

    # Delete reviews for all results in this run
    result_ids = (
        db.query(QATestRunResult.id)
        .filter(QATestRunResult.test_run_id == run_id)
        .all()
    )
    result_id_list = [r[0] for r in result_ids]

    if result_id_list:
        db.query(QATestRunReview).filter(
            QATestRunReview.test_run_result_id.in_(result_id_list)
        ).delete(synchronize_session=False)

    # Delete results
    db.query(QATestRunResult).filter(
        QATestRunResult.test_run_id == run_id
    ).delete(synchronize_session=False)

    # Delete the run itself
    db.delete(test_run)
    db.commit()

    logger.info("Deleted test run %s with %d results", run_id, len(result_id_list))


# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------

# Column headers for CSV/XLSX export (Requirement 9.3)
_EXPORT_HEADERS = [
    "departure_icao_aerodrome_code",
    "destination_aerodrome_code",
    "operational_icao_carrier_code",
    "flight_number",
    "icao_route",
    "resolved_waypoint_count",
    "fir_crossing_count",
    "unresolved_token_count",
    "resolved_waypoints",
    "fir_crossings",
    "unresolved_tokens",
    "error_message",
    "latest_verdict",
    "latest_reviewer_notes",
]


def _build_export_rows(run_id: str, db: Session) -> list[list]:
    """Query all results for a test run and build export rows.

    Each row contains the fields specified by Requirement 9.3: flight plan
    fields, counts, full JSONB lists as strings, error message, and the
    latest review verdict and notes.

    Args:
        run_id: UUID of the test run.
        db: SQLAlchemy database session.

    Returns:
        List of row lists (one per result), ready for CSV/XLSX writing.

    Raises:
        ValueError: If the test run does not exist.
    """
    test_run = db.query(QATestRun).filter(QATestRun.id == run_id).first()
    if not test_run:
        raise ValueError(f"Test run {run_id} not found")

    # Subquery: latest review per result (same pattern as get_test_run_detail)
    latest_review_sq = (
        db.query(
            QATestRunReview.test_run_result_id,
            QATestRunReview.verdict,
            QATestRunReview.reviewer_notes,
            func.row_number()
            .over(
                partition_by=QATestRunReview.test_run_result_id,
                order_by=QATestRunReview.reviewed_at.desc(),
            )
            .label("rn"),
        )
        .subquery("latest_review")
    )

    rows_query = (
        db.query(
            QATestRunResult,
            QAFlightPlan,
            latest_review_sq.c.verdict.label("latest_verdict"),
            latest_review_sq.c.reviewer_notes.label("latest_reviewer_notes"),
        )
        .join(QAFlightPlan, QATestRunResult.flight_plan_id == QAFlightPlan.id)
        .outerjoin(
            latest_review_sq,
            (latest_review_sq.c.test_run_result_id == QATestRunResult.id)
            & (latest_review_sq.c.rn == 1),
        )
        .filter(QATestRunResult.test_run_id == test_run.id)
    )

    export_rows: list[list] = []
    for result, fp, lat_verdict, lat_notes in rows_query.all():
        resolved_wp = result.resolved_waypoints or []
        fir_cx = result.fir_crossings or []
        unresolved_tk = result.unresolved_tokens or []

        export_rows.append([
            fp.departure_icao_aerodrome_code or "",
            fp.destination_aerodrome_code or "",
            fp.operational_icao_carrier_code or "",
            fp.flight_number or "",
            fp.icao_route or "",
            len(resolved_wp),
            len(fir_cx),
            len(unresolved_tk),
            json.dumps(resolved_wp),
            json.dumps(fir_cx),
            json.dumps(unresolved_tk),
            result.error_message or "",
            lat_verdict or "",
            lat_notes or "",
        ])

    return export_rows


def export_results_csv(run_id: str, db: Session) -> io.StringIO:
    """Export test run results as CSV. Returns a StringIO object.

    Includes all required fields per Requirement 9.3: flight plan fields,
    counts, full JSONB lists, error message, latest review verdict and notes.

    Args:
        run_id: UUID of the test run to export.
        db: SQLAlchemy database session.

    Returns:
        ``io.StringIO`` containing the CSV data.

    Raises:
        ValueError: If the test run does not exist.

    Validates Requirements: 9.1, 9.3
    """
    rows = _build_export_rows(run_id, db)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_EXPORT_HEADERS)
    writer.writerows(rows)
    output.seek(0)
    return output


def export_results_xlsx(run_id: str, db: Session) -> io.BytesIO:
    """Export test run results as Excel. Returns a BytesIO object.

    Includes all required fields per Requirement 9.3: flight plan fields,
    counts, full JSONB lists, error message, latest review verdict and notes.

    Args:
        run_id: UUID of the test run to export.
        db: SQLAlchemy database session.

    Returns:
        ``io.BytesIO`` containing the ``.xlsx`` data.

    Raises:
        ValueError: If the test run does not exist.

    Validates Requirements: 9.2, 9.3
    """
    from openpyxl import Workbook

    rows = _build_export_rows(run_id, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Run Results"
    ws.append(_EXPORT_HEADERS)
    for row in rows:
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
